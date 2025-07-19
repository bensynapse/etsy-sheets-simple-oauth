"""
Data management for CSV/Excel operations.
Handles template creation and data I/O.
"""

import pandas as pd
from typing import List, Dict, Optional, Any, Union
from datetime import datetime
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class DataManager:
    """Manages CSV/Excel data operations matching GAS implementation."""
    
    def __init__(self):
        """Initialize data manager."""
        self.template_columns = [
            'Title*', 
            'Description*', 
            'Price*', 
            'Quantity*', 
            'SKU', 
            'Tags (comma separated)', 
            'Materials (comma separated)',
            'Who Made*', 
            'When Made*', 
            'Taxonomy ID*', 
            'Image URLs (comma separated)', 
            'Status', 
            'Result', 
            'Delete?'
        ]
        
    def create_product_template(self, filename: str = 'product_upload_template.csv') -> pd.DataFrame:
        """Create product upload template (createProductTemplate).
        
        Matches the GAS implementation with 10 sample products.
        
        Args:
            filename: Output filename
            
        Returns:
            Template DataFrame
        """
        # Sample products matching GAS implementation exactly
        sample_products = [
            # 1. Ceramic Mug
            {
                'Title*': 'Handmade Blue Ceramic Coffee Mug',
                'Description*': 'Beautiful handcrafted ceramic mug with a stunning blue glaze. Perfect for your morning coffee or evening tea. Dishwasher and microwave safe. Holds 12oz of your favorite beverage.',
                'Price*': 24.99,
                'Quantity*': 15,
                'SKU': 'MUG-BLUE-001',
                'Tags (comma separated)': 'ceramic, handmade, mug, coffee, blue, kitchen, dishwasher safe',
                'Materials (comma separated)': 'ceramic, glaze',
                'Who Made*': 'i_did',
                'When Made*': '2020_2024',
                'Taxonomy ID*': 1633,
                'Image URLs (comma separated)': 'https://picsum.photos/800/600?random=1',
                'Status': '',
                'Result': '',
                'Delete?': ''
            },
            # 2. Wooden Cutting Board
            {
                'Title*': 'Rustic Wood Cutting Board - Walnut',
                'Description*': 'Handcrafted walnut cutting board perfect for your kitchen. Each board is unique with natural wood grain patterns. Treated with food-safe mineral oil. Makes a great wedding gift!',
                'Price*': 45.00,
                'Quantity*': 8,
                'SKU': 'BOARD-WAL-002',
                'Tags (comma separated)': 'cutting board, wood, walnut, kitchen, handmade, rustic, wedding gift',
                'Materials (comma separated)': 'walnut, mineral oil',
                'Who Made*': 'collective',
                'When Made*': 'made_to_order',
                'Taxonomy ID*': 1625,
                'Image URLs (comma separated)': 'https://picsum.photos/800/600?random=2',
                'Status': '',
                'Result': '',
                'Delete?': ''
            },
            # 3. Knitted Scarf
            {
                'Title*': 'Cozy Hand Knit Winter Scarf - Burgundy',
                'Description*': 'Soft and warm hand-knitted scarf in beautiful burgundy color. Made from premium merino wool. Perfect for cold winter days. Measures 70 inches long by 8 inches wide.',
                'Price*': 38.50,
                'Quantity*': 12,
                'SKU': 'SCARF-BURG-003',
                'Tags (comma separated)': 'scarf, knitted, winter, burgundy, merino wool, handmade, cozy, warm',
                'Materials (comma separated)': 'merino wool',
                'Who Made*': 'i_did',
                'When Made*': '2020_2024',
                'Taxonomy ID*': 166,
                'Image URLs (comma separated)': 'https://picsum.photos/800/600?random=3',
                'Status': '',
                'Result': '',
                'Delete?': ''
            },
            # 4. Vintage Jewelry
            {
                'Title*': 'Vintage Sterling Silver Locket Necklace',
                'Description*': 'Beautiful vintage sterling silver locket from the 1960s. Features intricate floral engravings. Chain length 18 inches. Perfect condition, professionally cleaned and polished.',
                'Price*': 125.00,
                'Quantity*': 3,
                'SKU': 'LOCKET-VINT-004',
                'Tags (comma separated)': 'vintage, locket, sterling silver, necklace, jewelry, 1960s, antique',
                'Materials (comma separated)': 'sterling silver',
                'Who Made*': 'someone_else',
                'When Made*': 'before_2005',
                'Taxonomy ID*': 1875,
                'Image URLs (comma separated)': 'https://picsum.photos/800/600?random=4',
                'Status': '',
                'Result': '',
                'Delete?': ''
            },
            # 5. Digital Art Print
            {
                'Title*': 'Modern Abstract Wall Art Print - Instant Download',
                'Description*': 'Instant download digital art print. Modern abstract design perfect for home or office decor. High resolution 300 DPI. Includes 5 sizes: 8x10, 11x14, 16x20, 18x24, 24x30 inches.',
                'Price*': 12.99,
                'Quantity*': 999,
                'SKU': 'PRINT-ABST-005',
                'Tags (comma separated)': 'digital print, wall art, abstract, instant download, printable, modern art',
                'Materials (comma separated)': 'digital file',
                'Who Made*': 'i_did',
                'When Made*': '2020_2024',
                'Taxonomy ID*': 2079,
                'Image URLs (comma separated)': 'https://picsum.photos/800/600?random=5',
                'Status': '',
                'Result': '',
                'Delete?': ''
            },
            # 6. Handmade Soap
            {
                'Title*': 'Lavender Goat Milk Soap - Natural & Organic',
                'Description*': 'Luxurious handmade soap with lavender essential oil and goat milk. Gentle on sensitive skin. Each bar is approximately 4oz. Wrapped in eco-friendly packaging.',
                'Price*': 8.99,
                'Quantity*': 25,
                'SKU': 'SOAP-LAV-006',
                'Tags (comma separated)': 'soap, handmade, lavender, goat milk, natural, organic, bath, skincare',
                'Materials (comma separated)': 'goat milk, lavender oil, coconut oil, shea butter',
                'Who Made*': 'i_did',
                'When Made*': '2020_2024',
                'Taxonomy ID*': 316,
                'Image URLs (comma separated)': 'https://picsum.photos/800/600?random=6',
                'Status': '',
                'Result': '',
                'Delete?': ''
            },
            # 7. Leather Wallet
            {
                'Title*': 'Personalized Leather Bifold Wallet',
                'Description*': 'Handcrafted genuine leather bifold wallet. Can be personalized with initials. Features 6 card slots, 2 bill compartments. Available in brown or black. Makes a perfect gift!',
                'Price*': 55.00,
                'Quantity*': 20,
                'SKU': 'WALLET-LEATH-007',
                'Tags (comma separated)': 'wallet, leather, personalized, mens wallet, bifold, gift for him',
                'Materials (comma separated)': 'genuine leather, brass hardware',
                'Who Made*': 'i_did',
                'When Made*': 'made_to_order',
                'Taxonomy ID*': 1811,
                'Image URLs (comma separated)': 'https://picsum.photos/800/600?random=7',
                'Status': '',
                'Result': '',
                'Delete?': ''
            },
            # 8. Plant Pot
            {
                'Title*': 'Minimalist Concrete Planter with Drainage',
                'Description*': 'Modern concrete planter perfect for succulents or small plants. Features drainage hole and cork pad. Hand-poured and sealed. Available in 4 inch diameter.',
                'Price*': 18.50,
                'Quantity*': 30,
                'SKU': 'PLANTER-CONC-008',
                'Tags (comma separated)': 'planter, concrete, succulent, minimalist, modern, home decor, plants',
                'Materials (comma separated)': 'concrete, cork',
                'Who Made*': 'i_did',
                'When Made*': '2020_2024',
                'Taxonomy ID*': 1633,
                'Image URLs (comma separated)': 'https://picsum.photos/800/600?random=8',
                'Status': '',
                'Result': '',
                'Delete?': ''
            },
            # 9. Vintage Book
            {
                'Title*': 'Rare First Edition Poetry Book 1952',
                'Description*': 'First edition poetry collection from 1952. Good vintage condition with some wear to dust jacket. All pages intact. A wonderful addition to any book collection.',
                'Price*': 75.00,
                'Quantity*': 1,
                'SKU': 'BOOK-POET-009',
                'Tags (comma separated)': 'vintage book, poetry, first edition, 1952, collectible, rare book',
                'Materials (comma separated)': 'paper, cloth binding',
                'Who Made*': 'someone_else',
                'When Made*': 'before_2005',
                'Taxonomy ID*': 161,
                'Image URLs (comma separated)': 'https://picsum.photos/800/600?random=9',
                'Status': '',
                'Result': '',
                'Delete?': ''
            },
            # 10. Candle Set
            {
                'Title*': 'Soy Candle Gift Set - Spring Collection',
                'Description*': 'Set of 3 hand-poured soy candles in spring scents: Cherry Blossom, Fresh Linen, and Garden Mint. Each candle burns for 20+ hours. Comes in beautiful gift box.',
                'Price*': 42.00,
                'Quantity*': 18,
                'SKU': 'CANDLE-SET-010',
                'Tags (comma separated)': 'candles, soy candle, gift set, spring, aromatherapy, home fragrance',
                'Materials (comma separated)': 'soy wax, cotton wick, essential oils',
                'Who Made*': 'i_did',
                'When Made*': '2020_2024',
                'Taxonomy ID*': 337,
                'Image URLs (comma separated)': 'https://picsum.photos/800/600?random=10',
                'Status': '',
                'Result': '',
                'Delete?': ''
            }
        ]
        
        # Create DataFrame
        df = pd.DataFrame(sample_products)
        
        # Add instructions at the bottom
        instructions = [
            {'Title*': '', 'Description*': '', 'Price*': '', 'Quantity*': ''},
            {'Title*': 'INSTRUCTIONS:', 'Description*': '', 'Price*': '', 'Quantity*': ''},
            {'Title*': '1. These are test products with real image URLs from Unsplash', 'Description*': '', 'Price*': '', 'Quantity*': ''},
            {'Title*': '2. Review and modify the products as needed before uploading', 'Description*': '', 'Price*': '', 'Quantity*': ''},
            {'Title*': '3. Who Made options: i_did, someone_else, collective', 'Description*': '', 'Price*': '', 'Quantity*': ''},
            {'Title*': '4. When Made options: made_to_order, 2020_2025, 2010_2019, 2006_2009, before_2005', 'Description*': '', 'Price*': '', 'Quantity*': ''},
            {'Title*': '5. Delete these instruction rows before uploading', 'Description*': '', 'Price*': '', 'Quantity*': ''},
            {'Title*': '6. Check Status and Result columns after upload for success/errors', 'Description*': '', 'Price*': '', 'Quantity*': ''},
            {'Title*': '7. To delete listings, put "X" or "x" in the Delete? column and click Delete Marked Listings', 'Description*': '', 'Price*': '', 'Quantity*': ''}
        ]
        
        # Fill empty columns for instructions
        for inst in instructions:
            for col in self.template_columns:
                if col not in inst:
                    inst[col] = ''
                    
        instructions_df = pd.DataFrame(instructions, columns=self.template_columns)
        
        # Combine data and instructions
        final_df = pd.concat([df, instructions_df], ignore_index=True)
        
        # Save to file
        final_df.to_csv(filename, index=False)
        
        logger.info(f"Created product template: {filename}")
        
        return final_df
        
    def read_product_data(self, file_path: Union[str, Path]) -> pd.DataFrame:
        """Read product data from CSV or Excel.
        
        Args:
            file_path: Path to file
            
        Returns:
            Product DataFrame
        """
        file_path = Path(file_path)
        
        if file_path.suffix.lower() == '.csv':
            df = pd.read_csv(file_path)
        elif file_path.suffix.lower() in ['.xlsx', '.xls']:
            df = pd.read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_path.suffix}")
            
        # Validate columns
        missing_cols = set(self.template_columns) - set(df.columns)
        if missing_cols:
            logger.warning(f"Missing columns: {missing_cols}")
            
        return df
        
    def update_product_status(self, df: pd.DataFrame, index: int, 
                            status: str, result: str = '') -> pd.DataFrame:
        """Update product status in DataFrame.
        
        Args:
            df: Product DataFrame
            index: Row index
            status: Status message
            result: Result message
            
        Returns:
            Updated DataFrame
        """
        df.loc[index, 'Status'] = status
        if result:
            df.loc[index, 'Result'] = result
        return df
        
    def save_results(self, df: pd.DataFrame, filename: str = 'upload_results.xlsx'):
        """Save results to Excel with formatting.
        
        Args:
            df: Results DataFrame
            filename: Output filename
        """
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Results', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Results']
            
            # Apply formatting
            from openpyxl.styles import PatternFill, Font
            
            # Header formatting
            header_fill = PatternFill(start_color='D3D3D3', 
                                    end_color='D3D3D3', 
                                    fill_type='solid')
            header_font = Font(bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                
            # Status column coloring
            status_col = None
            for idx, col in enumerate(worksheet[1]):
                if col.value == 'Status':
                    status_col = idx + 1
                    break
                    
            if status_col:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=status_col)
                    if '✓' in str(cell.value):
                        cell.fill = PatternFill(start_color='90EE90',
                                              end_color='90EE90',
                                              fill_type='solid')
                    elif '✗' in str(cell.value):
                        cell.fill = PatternFill(start_color='FFB6C1',
                                              end_color='FFB6C1',
                                              fill_type='solid')
                        
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        logger.info(f"Saved results to: {filename}")
        
    def export_summary(self, results: List[Dict[str, Any]], 
                      filename: str = 'upload_summary.xlsx'):
        """Export upload summary with statistics.
        
        Args:
            results: List of upload results
            filename: Output filename
        """
        summary_data = {
            'Metric': ['Total Products', 'Successful', 'Failed', 'Success Rate'],
            'Value': [
                len(results),
                sum(1 for r in results if r.get('success')),
                sum(1 for r in results if not r.get('success')),
                f"{sum(1 for r in results if r.get('success')) / len(results) * 100:.1f}%" if results else "0%"
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        results_df = pd.DataFrame(results)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            results_df.to_excel(writer, sheet_name='Details', index=False)
            
        logger.info(f"Exported summary to: {filename}")